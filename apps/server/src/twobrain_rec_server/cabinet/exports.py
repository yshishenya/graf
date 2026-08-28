from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import asdict, dataclass, replace
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from typing import Literal
from uuid import UUID

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from twobrain_rec_server.api.problems import ProblemDetail
from twobrain_rec_server.api.schemas import CONTENT_EXPORT_FORMATS_BY_SCOPE
from twobrain_rec_server.cabinet.speakers import speaker_names_for_result
from twobrain_rec_server.cabinet.view_models import source_role_label
from twobrain_rec_server.db.models import (
    DiarizationSegment,
    MediaRevision,
    Meeting,
    MeetingOutcomeItem,
    MeetingOutcomeSet,
    MeetingSpeakerName,
    ProcessingResult,
    TranscriptSegment,
)
from twobrain_rec_server.domain.speaker_turns import (
    CanonicalSpeakerModel,
    canonical_speaker_model,
    canonical_speech_available,
)
from twobrain_rec_server.domain.statuses import (
    MediaRevisionStatus,
    ProcessingAvailabilityStatus,
    ProcessingResultStatus,
)
from twobrain_rec_server.processing import store as processing_store

ExportScope = Literal["transcript", "summary", "combined"]
ExportFormat = Literal["txt", "md", "csv", "xlsx", "json", "srt", "vtt"]
AttributionState = Literal["confirmed", "unconfirmed", "unknown", "mixed", "uncertain"]

SCHEMA_VERSION = "graf.transcript-export.v2"
RENDERER_VERSION = "export-v1"
TURN_POLICY_VERSION = "canonical-provider-turns-v4"
UNKNOWN_SPEAKER_LABEL = "Спикер не определён"
FORMAT_COMPATIBILITY = CONTENT_EXPORT_FORMATS_BY_SCOPE
MEDIA_TYPES: dict[ExportFormat, str] = {
    "txt": "text/plain; charset=utf-8",
    "md": "text/markdown; charset=utf-8",
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "json": "application/json; charset=utf-8",
    "srt": "application/x-subrip; charset=utf-8",
    "vtt": "text/vtt; charset=utf-8",
}
CSV_COLUMNS = (
    "sequence",
    "turn_id",
    "start_ms",
    "end_ms",
    "start_time",
    "end_time",
    "speaker_key",
    "speaker_label",
    "provider_speaker_key",
    "attribution_state",
    "result_state",
    "source_role",
    "text",
    "overlap",
    "source_segment_ids",
    "processing_result_id",
    "turn_policy_version",
)
SUMMARY_CATEGORY_ORDER = (
    "summary",
    "key_points",
    "decisions",
    "action_items",
    "followups",
    "risks",
    "questions",
    "evidence",
)
SUMMARY_CATEGORY_LABELS = {
    "summary": "Саммари",
    "key_points": "Ключевые моменты",
    "decisions": "Решения",
    "action_items": "Задачи",
    "followups": "Следующие шаги",
    "risks": "Риски",
    "questions": "Вопросы",
    "evidence": "Основания",
}


@dataclass(frozen=True, slots=True)
class ExportSelection:
    content_scope: ExportScope
    format: ExportFormat
    processing_result_id: UUID
    outcome_set_id: UUID | None = None
    include_speaker_labels: bool = True
    include_timestamps: bool = True
    include_evidence: bool = True


@dataclass(frozen=True, slots=True)
class RawExportSegment:
    segment_id: str
    sequence: int
    start_ms: int
    end_ms: int
    text: str
    source_role: str
    source_role_original: str | None
    speaker_key: str
    speaker_label: str
    attribution_state: AttributionState
    timing_state: Literal["valid", "invalid"]
    omission_reason: str | None
    provider_speaker_key: str | None = None
    result_state: Literal["accepted", "degraded_provider_result"] = "accepted"


@dataclass(frozen=True, slots=True)
class CanonicalExportTurn:
    turn_id: str
    sequence: int
    start_ms: int
    end_ms: int
    text: str
    speaker_key: str
    speaker_label: str
    attribution_state: AttributionState
    source_role: str
    source_segment_ids: tuple[str, ...]
    overlap: bool
    provider_speaker_key: str | None = None
    result_state: Literal["accepted", "degraded_provider_result"] = "accepted"
    timing_state: Literal["valid", "invalid"] = "valid"


@dataclass(frozen=True, slots=True)
class SummaryExportItem:
    category: str
    sequence: int
    state: str
    text: str | None
    owner: str | None
    due_date: str | None
    truth_label: str
    source_references: tuple[dict[str, object], ...]
    evidence_turn_ids: tuple[str, ...]
    unresolved_references: tuple[dict[str, object], ...]


@dataclass(frozen=True, slots=True)
class SummaryExportRevision:
    outcome_set_id: str
    processing_result_id: str
    revision_token: str
    status: str
    category_states: dict[str, str]
    source_kind: str
    generator_kind: str
    generator_version: str
    content_hash: str | None
    items: tuple[SummaryExportItem, ...]


@dataclass(frozen=True, slots=True)
class ExportSnapshot:
    selection: ExportSelection
    meeting_id: str
    meeting_title: str
    language: str | None
    duration_seconds: int
    processing_result_id: str
    processing_result_version: int
    media_revision_id: str | None
    raw_segments: tuple[RawExportSegment, ...]
    canonical_turns: tuple[CanonicalExportTurn, ...]
    summary: SummaryExportRevision | None
    attribution_result_state: Literal["accepted", "degraded_provider_result"] = "accepted"
    attribution_reason_codes: tuple[str, ...] = ()
    schema_version: str = SCHEMA_VERSION
    renderer_version: str = RENDERER_VERSION
    turn_policy_version: str = TURN_POLICY_VERSION


@dataclass(frozen=True, slots=True)
class GeneratedContentExport:
    filename: str
    media_type: str
    body: bytes

    @property
    def byte_length(self) -> int:
        return len(self.body)


def validate_export_selection(selection: ExportSelection) -> None:
    if selection.format not in FORMAT_COMPATIBILITY[selection.content_scope]:
        raise ProblemDetail(
            status=422,
            code="unsupported_export_combination",
            title="Unsupported export combination",
        )
    summary_requested = selection.content_scope in {"summary", "combined"}
    if summary_requested != (selection.outcome_set_id is not None):
        raise ProblemDetail(
            status=422,
            code="invalid_export_selection",
            title="Invalid export selection",
        )


async def build_export_snapshot(
    db: AsyncSession,
    *,
    meeting: Meeting,
    result: ProcessingResult,
    selection: ExportSelection,
) -> ExportSnapshot:
    validate_export_selection(selection)
    selection = _effective_export_selection(selection)
    transcript_requested = selection.content_scope in {"transcript", "combined"}
    current_revision = await db.scalar(
        select(MediaRevision)
        .where(
            MediaRevision.workspace_id == meeting.workspace_id,
            MediaRevision.meeting_id == meeting.id,
            MediaRevision.status == MediaRevisionStatus.ACCEPTED.value,
            MediaRevision.immutable.is_(True),
        )
        .order_by(MediaRevision.revision_number.desc(), MediaRevision.updated_at.desc())
    )
    effective_result = await processing_store.latest_processing_result(
        db,
        workspace_id=meeting.workspace_id,
        meeting_id=meeting.id,
        media_revision_id=current_revision.id if current_revision is not None else None,
    )
    if (
        result.id != selection.processing_result_id
        or result.workspace_id != meeting.workspace_id
        or result.meeting_id != meeting.id
        or effective_result is None
        or effective_result.id != result.id
        or result.status != ProcessingResultStatus.IMPORTED.value
        or (
            transcript_requested
            and (
                not canonical_speech_available(result)
                or
                result.transcript_status != ProcessingAvailabilityStatus.AVAILABLE.value
                or result.segment_count <= 0
                or result.diarization_status != ProcessingAvailabilityStatus.AVAILABLE.value
                or result.diarization_segment_count <= 0
            )
        )
    ):
        raise _stale_selection()

    transcript_rows = list(
        (
            await db.scalars(
                select(TranscriptSegment)
                .where(
                    TranscriptSegment.workspace_id == meeting.workspace_id,
                    TranscriptSegment.meeting_id == meeting.id,
                    TranscriptSegment.processing_result_id == result.id,
                )
                .order_by(TranscriptSegment.sequence.asc(), TranscriptSegment.start_seconds.asc())
            )
        ).all()
    )
    diarization_rows = list(
        (
            await db.scalars(
                select(DiarizationSegment)
                .where(
                    DiarizationSegment.workspace_id == meeting.workspace_id,
                    DiarizationSegment.meeting_id == meeting.id,
                    DiarizationSegment.processing_result_id == result.id,
                )
                .order_by(DiarizationSegment.start_seconds.asc(), DiarizationSegment.sequence.asc())
            )
        ).all()
    )
    transcript_rows_match = _same_processing_result_rows(
        transcript_rows, expected_result_id=result.id
    )
    diarization_rows_match = _same_processing_result_rows(
        diarization_rows, expected_result_id=result.id
    )
    transcript_visible = bool(
        result.transcript_status == ProcessingAvailabilityStatus.AVAILABLE.value
        and result.segment_count > 0
        and result.diarization_status == ProcessingAvailabilityStatus.AVAILABLE.value
        and result.diarization_segment_count > 0
        and transcript_rows
        and transcript_rows_match
        and diarization_rows
        and diarization_rows_match
    )
    if transcript_requested and not transcript_visible:
        raise ProblemDetail(status=409, code="export_unavailable", title="Export unavailable")
    # Summary is an independent value stream.  It may be exported while the
    # transcript remains an internal transcript-only/diarization-pending result;
    # do not manufacture transcript evidence in that case.
    if not transcript_visible:
        transcript_rows = []
        diarization_rows = []
    speaker_names = speaker_names_for_result(
        (
            await db.scalars(
                select(MeetingSpeakerName).where(
                    MeetingSpeakerName.workspace_id == meeting.workspace_id,
                    MeetingSpeakerName.meeting_id == meeting.id,
                )
            )
        ).all(),
        result_imported_at=result.imported_at,
    )
    model = canonical_speaker_model(
        transcript_rows,
        diarization_rows,
        processing_result_id=result.id,
        speaker_names=speaker_names,
        source_result_hash=result.source_result_hash,
    )
    raw_segments = canonical_raw_segments(
        transcript_rows,
        result_state=model.result_state,
    )
    turns = _export_turns_from_model(model)
    summary = await _load_summary_revision(
        db,
        meeting=meeting,
        result=result,
        outcome_set_id=selection.outcome_set_id,
        turns=turns,
    )
    return ExportSnapshot(
        selection=selection,
        meeting_id=str(meeting.id),
        meeting_title=_safe_title(meeting.title),
        language=result.language,
        duration_seconds=max(meeting.duration_seconds, 0),
        processing_result_id=str(result.id),
        processing_result_version=result.result_version,
        media_revision_id=str(result.media_revision_id) if result.media_revision_id else None,
        raw_segments=raw_segments,
        canonical_turns=turns,
        summary=summary,
        attribution_result_state=model.result_state,
        attribution_reason_codes=model.diagnostics.reason_codes,
    )


def canonical_export_turns(
    transcript_rows: list[TranscriptSegment],
    *,
    diarization_rows: list[DiarizationSegment],
    processing_result_id: UUID,
    speaker_names: dict[str, str] | None = None,
) -> tuple[CanonicalExportTurn, ...]:
    model = canonical_speaker_model(
        transcript_rows,
        diarization_rows,
        processing_result_id=processing_result_id,
        speaker_names=speaker_names,
    )
    return _export_turns_from_model(model)


def _export_turns_from_model(model: CanonicalSpeakerModel) -> tuple[CanonicalExportTurn, ...]:
    return tuple(
        CanonicalExportTurn(
            turn_id=turn.turn_id,
            sequence=turn.sequence,
            start_ms=_milliseconds(turn.start_seconds),
            end_ms=_milliseconds(turn.end_seconds),
            text=turn.text,
            speaker_key=turn.speaker_key,
            speaker_label=turn.speaker_label,
            provider_speaker_key=turn.provider_speaker_key,
            attribution_state=turn.attribution_state,
            result_state=turn.result_state,
            source_role=source_role_label(turn.source_role),
            source_segment_ids=(turn.source_segment_id,),
            overlap=turn.overlap,
        )
        for turn in model.turns
    )


def _same_processing_result_rows(
    rows: list[TranscriptSegment] | list[DiarizationSegment],
    *,
    expected_result_id: UUID,
) -> bool:
    result_ids = {row.processing_result_id for row in rows}
    return bool(result_ids) and result_ids == {expected_result_id}


def canonical_raw_segments(
    transcript_rows: list[TranscriptSegment],
    *,
    result_state: Literal["accepted", "degraded_provider_result"] = "accepted",
) -> tuple[RawExportSegment, ...]:
    output: list[RawExportSegment] = []
    for row in transcript_rows:
        valid_timing = row.start_seconds >= 0 and row.end_seconds > row.start_seconds
        omission_reason = None
        if not valid_timing:
            omission_reason = "invalid_timing"
        elif not row.text.strip():
            omission_reason = "empty_text"
        output.append(
            RawExportSegment(
                segment_id=str(row.id),
                sequence=row.sequence,
                start_ms=_milliseconds(row.start_seconds),
                end_ms=_milliseconds(row.end_seconds),
                text=row.text,
                source_role=source_role_label(row.source_role),
                source_role_original=row.source_role_original,
                speaker_key=f"evidence:{row.processing_result_id.hex}",
                speaker_label=UNKNOWN_SPEAKER_LABEL,
                attribution_state="uncertain",
                result_state=result_state,
                timing_state="valid" if valid_timing else "invalid",
                omission_reason=omission_reason,
            )
        )
    return tuple(output)


async def _load_summary_revision(
    db: AsyncSession,
    *,
    meeting: Meeting,
    result: ProcessingResult,
    outcome_set_id: UUID | None,
    turns: tuple[CanonicalExportTurn, ...],
) -> SummaryExportRevision | None:
    if outcome_set_id is None:
        return None
    outcome_set = await db.scalar(
        select(MeetingOutcomeSet).where(
            MeetingOutcomeSet.id == outcome_set_id,
            MeetingOutcomeSet.workspace_id == meeting.workspace_id,
            MeetingOutcomeSet.meeting_id == meeting.id,
            MeetingOutcomeSet.processing_result_id == result.id,
            MeetingOutcomeSet.lifecycle_state == "active",
            MeetingOutcomeSet.revision_state == "accepted",
            MeetingOutcomeSet.id == meeting.current_outcome_set_id,
        )
    )
    if outcome_set is None:
        raise _stale_selection()
    rows = list(
        (
            await db.scalars(
                select(MeetingOutcomeItem)
                .where(
                    MeetingOutcomeItem.workspace_id == meeting.workspace_id,
                    MeetingOutcomeItem.meeting_id == meeting.id,
                    MeetingOutcomeItem.outcome_set_id == outcome_set.id,
                )
                .order_by(MeetingOutcomeItem.category.asc(), MeetingOutcomeItem.sequence.asc())
            )
        ).all()
    )
    category_order = {category: index for index, category in enumerate(SUMMARY_CATEGORY_ORDER)}
    rows.sort(
        key=lambda row: (
            category_order.get(row.category, len(category_order)),
            row.sequence,
        )
    )
    turn_by_segment = {
        segment_id: turn.turn_id for turn in turns for segment_id in turn.source_segment_ids
    }
    items: list[SummaryExportItem] = []
    for row in rows:
        references = tuple(ref for ref in row.source_refs_json if isinstance(ref, dict))
        resolved: list[str] = []
        unresolved: list[dict[str, object]] = []
        for reference in references:
            turn_ids = _reference_turn_ids(reference, turns, turn_by_segment)
            if not turn_ids:
                unresolved.append(reference)
                continue
            for turn_id in turn_ids:
                if turn_id not in resolved:
                    resolved.append(turn_id)
        items.append(
            SummaryExportItem(
                category=row.category,
                sequence=row.sequence,
                state=row.state,
                text=row.text,
                owner=row.owner_text,
                due_date=row.due_date_text,
                truth_label=row.truth_label,
                source_references=references,
                evidence_turn_ids=tuple(resolved),
                unresolved_references=tuple(unresolved),
            )
        )
    category_states = {
        category: str(getattr(outcome_set, f"{category}_state"))
        for category in SUMMARY_CATEGORY_ORDER
    }
    revision_token = ":".join(
        (
            str(outcome_set.id),
            outcome_set.content_hash or "no-content-hash",
            outcome_set.generator_version,
        )
    )
    return SummaryExportRevision(
        outcome_set_id=str(outcome_set.id),
        processing_result_id=str(result.id),
        revision_token=revision_token,
        status=outcome_set.status,
        category_states=category_states,
        source_kind=outcome_set.source_kind,
        generator_kind=outcome_set.generator_kind,
        generator_version=outcome_set.generator_version,
        content_hash=outcome_set.content_hash,
        items=tuple(items),
    )


def _reference_turn_ids(
    reference: dict[str, object],
    turns: tuple[CanonicalExportTurn, ...],
    turn_by_segment: dict[str, str],
) -> tuple[str, ...]:
    raw_id = reference.get("transcript_segment_id")
    direct = turn_by_segment.get(str(raw_id)) if raw_id is not None else None
    if direct is not None:
        return (direct,)
    try:
        start_ms = _milliseconds(Decimal(str(reference["start_seconds"])))
        end_ms = _milliseconds(Decimal(str(reference["end_seconds"])))
    except (KeyError, TypeError, ValueError, InvalidOperation):
        return ()
    if end_ms <= start_ms:
        return ()
    raw_role = reference.get("source_role")
    role = source_role_label(str(raw_role)) if raw_role else None
    return tuple(
        turn.turn_id
        for turn in turns
        if (role is None or turn.source_role == role)
        and min(end_ms, turn.end_ms) > max(start_ms, turn.start_ms)
    )


def render_content_export(snapshot: ExportSnapshot) -> GeneratedContentExport:
    snapshot = replace(
        snapshot,
        selection=_effective_export_selection(snapshot.selection),
    )
    if snapshot.selection.format in {"srt", "vtt"} and any(
        turn.end_ms <= turn.start_ms for turn in snapshot.canonical_turns
    ):
        raise ProblemDetail(
            status=409,
            code="subtitle_timing_unavailable",
            title="Subtitle timing unavailable",
        )
    renderer = {
        "txt": _render_txt,
        "md": _render_markdown,
        "csv": _render_csv,
        "xlsx": _render_xlsx,
        "json": _render_json,
        "srt": _render_srt,
        "vtt": _render_vtt,
    }[snapshot.selection.format]
    try:
        body = renderer(snapshot)
    except ProblemDetail:
        raise
    except Exception as exc:
        raise ProblemDetail(
            status=503,
            code="export_generation_failed",
            title="Export generation failed",
        ) from exc
    filename = (
        f"graf-meeting-{snapshot.meeting_id[:8]}-"
        f"{snapshot.selection.content_scope}-r{snapshot.processing_result_version}."
        f"{snapshot.selection.format}"
    )
    return GeneratedContentExport(
        filename=filename,
        media_type=MEDIA_TYPES[snapshot.selection.format],
        body=body,
    )


def _render_txt(snapshot: ExportSnapshot) -> bytes:
    lines = [
        snapshot.meeting_title,
        f"Состав: {_scope_label(snapshot.selection.content_scope)}",
        f"Ревизия транскрипта: {snapshot.processing_result_version}",
        f"Язык: {snapshot.language or 'не указан'}",
        f"Длительность: {_human_time(snapshot.duration_seconds * 1000)}",
        f"Разделение по спикерам: {_attribution_status_label(snapshot)}",
        "",
    ]
    if snapshot.selection.content_scope in {"transcript", "combined"}:
        if snapshot.selection.content_scope == "combined":
            lines.extend(("Транскрипт", "===========", ""))
        lines.extend(_human_transcript_lines(snapshot, markdown=False))
    if snapshot.selection.content_scope in {"summary", "combined"}:
        if snapshot.selection.content_scope == "combined":
            lines.extend(("", "Саммари", "=======", ""))
        lines.extend(_summary_lines(snapshot, markdown=False))
    return ("\n".join(lines).rstrip() + "\n").encode("utf-8")


def _render_markdown(snapshot: ExportSnapshot) -> bytes:
    lines = [
        f"# {_markdown_escape(snapshot.meeting_title)}",
        "",
        f"- Состав: {_scope_label(snapshot.selection.content_scope)}",
        f"- Ревизия транскрипта: {snapshot.processing_result_version}",
        f"- Язык: {_markdown_escape(snapshot.language or 'не указан')}",
        f"- Длительность: {_human_time(snapshot.duration_seconds * 1000)}",
        f"- Разделение по спикерам: {_markdown_escape(_attribution_status_label(snapshot))}",
        "",
    ]
    if snapshot.selection.content_scope in {"transcript", "combined"}:
        if snapshot.selection.content_scope == "combined":
            lines.extend(("## Транскрипт", ""))
        lines.extend(_human_transcript_lines(snapshot, markdown=True))
    if snapshot.selection.content_scope in {"summary", "combined"}:
        if snapshot.selection.content_scope == "combined":
            lines.extend(("", "## Саммари", ""))
        lines.extend(_summary_lines(snapshot, markdown=True))
    return ("\n".join(lines).rstrip() + "\n").encode("utf-8")


def _human_transcript_lines(snapshot: ExportSnapshot, *, markdown: bool) -> list[str]:
    lines: list[str] = []
    for group in human_display_groups(
        snapshot.canonical_turns,
        raw_segments=snapshot.raw_segments,
    ):
        first = group[0]
        heading = first.speaker_label if snapshot.selection.include_speaker_labels else "Реплики"
        if markdown:
            lines.extend((f"### {_markdown_escape(heading)}", ""))
        else:
            lines.append(heading)
        for turn in group:
            timestamp = (
                f"[{_human_time(turn.start_ms)}] " if snapshot.selection.include_timestamps else ""
            )
            text = _markdown_escape(turn.text) if markdown else turn.text
            lines.append(f"{timestamp}{text}")
        lines.append("")
    return lines


def human_display_groups(
    turns: tuple[CanonicalExportTurn, ...],
    *,
    raw_segments: tuple[RawExportSegment, ...] = (),
) -> tuple[tuple[CanonicalExportTurn, ...], ...]:
    source_boundaries = {
        row.segment_id: row.source_role_original or row.source_role for row in raw_segments
    }

    def source_boundary(turn: CanonicalExportTurn) -> str:
        return source_boundaries.get(turn.source_segment_ids[0], turn.source_role)

    groups: list[list[CanonicalExportTurn]] = []
    for turn in turns:
        previous = groups[-1][-1] if groups else None
        if (
            previous is not None
            and turn.attribution_state == "confirmed"
            and previous.attribution_state == "confirmed"
            and turn.speaker_key == previous.speaker_key
            and source_boundary(turn) == source_boundary(previous)
            and not turn.overlap
            and not previous.overlap
            and 0 <= turn.start_ms - previous.end_ms <= 3000
        ):
            groups[-1].append(turn)
        else:
            groups.append([turn])
    return tuple(tuple(group) for group in groups)


def _summary_lines(snapshot: ExportSnapshot, *, markdown: bool) -> list[str]:
    summary = snapshot.summary
    if summary is None:
        return ["Сохраненное саммари недоступно."]
    lines = [
        (
            f"Статус сохраненной ревизии: {_markdown_escape(summary.status)}"
            if markdown
            else f"Статус сохраненной ревизии: {summary.status}"
        ),
        (
            "Источник сохраненной ревизии: " + _markdown_escape(summary.source_kind)
            if markdown
            else f"Источник сохраненной ревизии: {summary.source_kind}"
        ),
        (
            "Генератор: "
            + _markdown_escape(f"{summary.generator_kind} {summary.generator_version}")
            if markdown
            else f"Генератор: {summary.generator_kind} {summary.generator_version}"
        ),
        "",
    ]
    by_category = {
        category: [item for item in summary.items if item.category == category]
        for category in SUMMARY_CATEGORY_ORDER
    }
    for category in SUMMARY_CATEGORY_ORDER:
        label = SUMMARY_CATEGORY_LABELS[category]
        state = summary.category_states[category]
        lines.append(f"### {label}" if markdown else label)
        if not markdown:
            lines.append("-" * len(label))
        items = by_category[category]
        if not items:
            lines.append(f"Состояние: {state}")
        for item in items:
            text = item.text or f"Состояние: {item.state}"
            text = _markdown_escape(text) if markdown else text
            lines.append(f"- {text}")
            if item.owner:
                owner = _markdown_escape(item.owner) if markdown else item.owner
                lines.append(f"  Ответственный: {owner}")
            if item.due_date:
                due = _markdown_escape(item.due_date) if markdown else item.due_date
                lines.append(f"  Срок: {due}")
            if snapshot.selection.include_evidence and item.evidence_turn_ids:
                lines.append(f"  Основания: {', '.join(item.evidence_turn_ids)}")
        lines.append("")
    return lines


def _render_csv(snapshot: ExportSnapshot) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=CSV_COLUMNS, lineterminator="\r\n")
    writer.writeheader()
    for turn in snapshot.canonical_turns:
        writer.writerow(_turn_row(snapshot, turn, safe_cells=True))
    return b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")


def _render_json(snapshot: ExportSnapshot) -> bytes:
    transcript = None
    if snapshot.selection.content_scope in {"transcript", "combined"}:
        transcript = {
            "status": snapshot.attribution_result_state,
            "reason_codes": snapshot.attribution_reason_codes,
            "raw_segments": [asdict(row) for row in snapshot.raw_segments],
            "canonical_turns": [asdict(turn) for turn in snapshot.canonical_turns],
        }
    summary = None
    if snapshot.selection.content_scope in {"summary", "combined"} and snapshot.summary:
        summary = asdict(snapshot.summary)
        if not snapshot.selection.include_evidence:
            for item in summary["items"]:
                item["source_references"] = ()
                item["evidence_turn_ids"] = ()
                item["unresolved_references"] = ()
    payload = {
        "schema_version": snapshot.schema_version,
        "renderer_version": snapshot.renderer_version,
        "meeting": {
            "meeting_id": snapshot.meeting_id,
            "title": snapshot.meeting_title,
            "language": snapshot.language,
            "duration_seconds": snapshot.duration_seconds,
        },
        "selection": {
            "content_scope": snapshot.selection.content_scope,
            "format": snapshot.selection.format,
            "include_speaker_labels": snapshot.selection.include_speaker_labels,
            "include_timestamps": snapshot.selection.include_timestamps,
            "include_evidence": snapshot.selection.include_evidence,
        },
        "revisions": {
            "processing_result_id": snapshot.processing_result_id,
            "processing_result_version": snapshot.processing_result_version,
            "media_revision_id": snapshot.media_revision_id,
            "outcome_set_id": snapshot.summary.outcome_set_id if snapshot.summary else None,
        },
        "transcript": transcript,
        "summary": summary,
        "provenance": {
            "turn_policy_version": snapshot.turn_policy_version,
            "provider_neutral": True,
        },
    }
    return (
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")) + "\n"
    ).encode("utf-8")


def _render_srt(snapshot: ExportSnapshot) -> bytes:
    blocks = []
    for counter, turn in enumerate(snapshot.canonical_turns, start=1):
        text = _subtitle_literal(turn.text)
        if snapshot.selection.include_speaker_labels:
            text = f"{_subtitle_literal(turn.speaker_label)}: {text}"
        blocks.append(f"{counter}\n{_srt_time(turn.start_ms)} --> {_srt_time(turn.end_ms)}\n{text}")
    return ("\n\n".join(blocks) + ("\n" if blocks else "")).encode("utf-8")


def _render_vtt(snapshot: ExportSnapshot) -> bytes:
    blocks = []
    for turn in snapshot.canonical_turns:
        text = _subtitle_literal(turn.text)
        if snapshot.selection.include_speaker_labels:
            text = f"{_subtitle_literal(turn.speaker_label)}: {text}"
        blocks.append(f"{_vtt_time(turn.start_ms)} --> {_vtt_time(turn.end_ms)}\n{text}")
    body = "WEBVTT\n\n" + "\n\n".join(blocks)
    return (body + ("\n" if blocks else "")).encode("utf-8")


def _render_xlsx(snapshot: ExportSnapshot) -> bytes:
    workbook = Workbook(write_only=True)
    transcript = workbook.create_sheet("Transcript")
    summary_sheet = workbook.create_sheet("Summary")
    action_items = workbook.create_sheet("Action Items")
    metadata = workbook.create_sheet("Metadata")
    _configure_sheet(transcript, CSV_COLUMNS)
    if snapshot.selection.content_scope in {"transcript", "combined"}:
        for turn in snapshot.canonical_turns:
            row = _turn_row(snapshot, turn, safe_cells=True)
            _append_sheet_row(transcript, [row[column] for column in CSV_COLUMNS])
    else:
        status_row = {column: "" for column in CSV_COLUMNS}
        status_row["turn_id"] = "status:not_selected"
        status_row["text"] = "not_selected"
        _append_sheet_row(transcript, [status_row[column] for column in CSV_COLUMNS])
    summary_columns = (
        "category",
        "sequence",
        "state",
        "text",
        "truth_label",
        "evidence_turn_ids",
        "source_references",
        "unresolved_references",
    )
    _configure_sheet(summary_sheet, summary_columns)
    action_columns = (
        "sequence",
        "state",
        "text",
        "owner",
        "due_date",
        "truth_label",
        "evidence_turn_ids",
        "source_references",
        "unresolved_references",
    )
    _configure_sheet(action_items, action_columns)
    if snapshot.summary is None:
        _append_sheet_row(
            summary_sheet,
            ("status", "", "not_selected", "", "", "", "", ""),
        )
        _append_sheet_row(
            action_items,
            ("", "not_selected", "", "", "", "", "", "", ""),
        )
    else:
        items_by_category = {
            category: [item for item in snapshot.summary.items if item.category == category]
            for category in SUMMARY_CATEGORY_ORDER
        }
        for category in SUMMARY_CATEGORY_ORDER:
            items = items_by_category[category]
            if category == "action_items":
                if not items:
                    _append_sheet_row(
                        action_items,
                        (
                            "",
                            snapshot.summary.category_states[category],
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ),
                    )
                for item in items:
                    evidence = _xlsx_evidence_values(snapshot, item)
                    _append_sheet_row(
                        action_items,
                        (
                            item.sequence,
                            item.state,
                            item.text,
                            item.owner,
                            item.due_date,
                            item.truth_label,
                            *evidence,
                        ),
                    )
                continue
            if not items:
                _append_sheet_row(
                    summary_sheet,
                    (
                        category,
                        "",
                        snapshot.summary.category_states[category],
                        "",
                        "",
                        "",
                        "",
                        "",
                    ),
                )
            for item in items:
                evidence = _xlsx_evidence_values(snapshot, item)
                _append_sheet_row(
                    summary_sheet,
                    (
                        category,
                        item.sequence,
                        item.state,
                        item.text,
                        item.truth_label,
                        *evidence,
                    ),
                )
    _configure_sheet(metadata, ("key", "value"))
    metadata_rows = (
        ("schema_version", snapshot.schema_version),
        ("renderer_version", snapshot.renderer_version),
        ("turn_policy_version", snapshot.turn_policy_version),
        ("meeting_id", snapshot.meeting_id),
        ("processing_result_id", snapshot.processing_result_id),
        ("processing_result_version", snapshot.processing_result_version),
        ("media_revision_id", snapshot.media_revision_id or ""),
        ("outcome_set_id", snapshot.summary.outcome_set_id if snapshot.summary else ""),
        (
            "summary_revision_token",
            snapshot.summary.revision_token if snapshot.summary else "",
        ),
        ("summary_source_kind", snapshot.summary.source_kind if snapshot.summary else ""),
        ("summary_generator_kind", snapshot.summary.generator_kind if snapshot.summary else ""),
        (
            "summary_generator_version",
            snapshot.summary.generator_version if snapshot.summary else "",
        ),
        ("summary_content_hash", snapshot.summary.content_hash or "" if snapshot.summary else ""),
        (
            "summary_category_states",
            json.dumps(snapshot.summary.category_states, ensure_ascii=False, sort_keys=True)
            if snapshot.summary
            else "",
        ),
        ("content_scope", snapshot.selection.content_scope),
        ("language", snapshot.language or ""),
        ("duration_seconds", snapshot.duration_seconds),
        ("attribution_result_state", snapshot.attribution_result_state),
        (
            "attribution_reason_codes",
            json.dumps(snapshot.attribution_reason_codes, ensure_ascii=False),
        ),
        ("summary_status", snapshot.summary.status if snapshot.summary else "not_selected"),
    )
    for row in metadata_rows:
        _append_sheet_row(metadata, row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _configure_sheet(sheet: object, columns: tuple[str, ...]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    widths = {
        "text": 72,
        "source_references": 48,
        "unresolved_references": 48,
        "evidence_turn_ids": 42,
        "source_segment_ids": 42,
        "speaker_label": 24,
        "speaker_key": 28,
        "turn_id": 30,
        "processing_result_id": 38,
        "value": 64,
    }
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(column, 18)
    cells = []
    for value in columns:
        cell = WriteOnlyCell(sheet, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cells.append(cell)
    sheet.append(cells)


def _xlsx_evidence_values(
    snapshot: ExportSnapshot,
    item: SummaryExportItem,
) -> tuple[str, str, str]:
    evidence_turn_ids = item.evidence_turn_ids if snapshot.selection.include_evidence else ()
    source_references = item.source_references if snapshot.selection.include_evidence else ()
    unresolved = item.unresolved_references if snapshot.selection.include_evidence else ()
    return tuple(
        json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        for value in (evidence_turn_ids, source_references, unresolved)
    )


def _append_sheet_row(sheet: object, values: object) -> None:
    cells = []
    for value in values:
        if isinstance(value, str):
            value = _safe_spreadsheet_text(value)
        cell = WriteOnlyCell(sheet, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cells.append(cell)
    sheet.append(cells)


def _turn_row(
    snapshot: ExportSnapshot,
    turn: CanonicalExportTurn,
    *,
    safe_cells: bool,
) -> dict[str, object]:
    clean = _safe_spreadsheet_text if safe_cells else lambda value: value
    return {
        "sequence": turn.sequence,
        "turn_id": turn.turn_id,
        "start_ms": turn.start_ms,
        "end_ms": turn.end_ms,
        "start_time": _human_time(turn.start_ms),
        "end_time": _human_time(turn.end_ms),
        "speaker_key": clean(turn.speaker_key),
        "speaker_label": clean(turn.speaker_label),
        "provider_speaker_key": clean(turn.provider_speaker_key),
        "attribution_state": turn.attribution_state,
        "result_state": turn.result_state,
        "source_role": turn.source_role,
        "text": clean(turn.text),
        "overlap": turn.overlap,
        "source_segment_ids": json.dumps(turn.source_segment_ids, ensure_ascii=False),
        "processing_result_id": snapshot.processing_result_id,
        "turn_policy_version": snapshot.turn_policy_version,
    }


def _safe_spreadsheet_text(value: str | None) -> str:
    if value is None:
        return ""
    return "'" + value if value.lstrip().startswith(("=", "+", "-", "@")) else value


def _markdown_escape(value: str) -> str:
    value = value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return re.sub(r"([!\"#$%\'()*+,./:=?@\\\[\]^_`{|}~-])", r"\\\1", value)


def _subtitle_literal(value: str) -> str:
    return " ".join(value.replace("<", "&lt;").replace(">", "&gt;").split())


def _milliseconds(value: Decimal) -> int:
    return int((value * 1000).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _human_time(milliseconds: int) -> str:
    total_seconds, millis = divmod(max(milliseconds, 0), 1000)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}.{millis:03d}"
    return f"{minutes:02d}:{seconds:02d}.{millis:03d}"


def _srt_time(milliseconds: int) -> str:
    total_seconds, millis = divmod(max(milliseconds, 0), 1000)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d},{millis:03d}"


def _vtt_time(milliseconds: int) -> str:
    return _srt_time(milliseconds).replace(",", ".")


def _safe_title(title: str | None) -> str:
    clean = " ".join((title or "Встреча GRAF").split())
    return clean[:160] or "Встреча GRAF"


def _scope_label(scope: ExportScope) -> str:
    return {"transcript": "транскрипт", "summary": "саммари", "combined": "транскрипт и саммари"}[
        scope
    ]


def _attribution_status_label(snapshot: ExportSnapshot) -> str:
    if snapshot.attribution_result_state == "accepted":
        return "готово"
    confirmed = any(turn.attribution_state == "confirmed" for turn in snapshot.canonical_turns)
    unconfirmed = any(turn.attribution_state != "confirmed" for turn in snapshot.canonical_turns)
    if confirmed and unconfirmed:
        return "частично готово; фрагменты без имени отмечены как «Спикер не определён»"
    if confirmed:
        return "показано по доступным данным; текст сохранён"
    return "без имён; текст сохранён"


def _effective_export_selection(selection: ExportSelection) -> ExportSelection:
    structural = selection.format in {"csv", "xlsx", "json"}
    return replace(
        selection,
        include_speaker_labels=True if structural else selection.include_speaker_labels,
        include_timestamps=(
            True
            if structural or selection.format in {"srt", "vtt"}
            else selection.include_timestamps
        ),
        include_evidence=(
            False if selection.content_scope == "transcript" else selection.include_evidence
        ),
    )


def _stale_selection() -> ProblemDetail:
    return ProblemDetail(
        status=409,
        code="export_revision_stale",
        title="Export revision is stale",
    )
