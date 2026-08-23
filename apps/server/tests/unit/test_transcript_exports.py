from __future__ import annotations

import csv
import io
import json
from dataclasses import replace
from decimal import Decimal
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from tests.fixtures.cabinet_exports import SyntheticExportFixture
from twobrain_rec_server.api.problems import ProblemDetail
from twobrain_rec_server.cabinet.exports import (
    CSV_COLUMNS,
    SCHEMA_VERSION,
    TURN_POLICY_VERSION,
    CanonicalExportTurn,
    ExportSelection,
    ExportSnapshot,
    SummaryExportItem,
    SummaryExportRevision,
    _reference_turn_ids,
    canonical_export_turns,
    canonical_raw_segments,
    human_display_groups,
    render_content_export,
)
from twobrain_rec_server.db.models import DiarizationSegment, TranscriptSegment
from twobrain_rec_server.domain.speaker_turns import stable_speaker_key


def _snapshot(
    fixture: SyntheticExportFixture,
    *,
    format: str,
    scope: str = "transcript",
    include_speaker_labels: bool = True,
    include_timestamps: bool = True,
) -> ExportSnapshot:
    raw = canonical_raw_segments(list(fixture.transcript_rows))
    turns = canonical_export_turns(
        list(fixture.transcript_rows),
        diarization_rows=list(fixture.diarization_rows),
        processing_result_id=fixture.result_id,
        speaker_names={stable_speaker_key(fixture.result_id, "speaker-a"): "Анна"},
    )
    summary = None
    outcome_set_id = None
    if scope in {"summary", "combined"}:
        outcome_set_id = uuid4()
        summary = SummaryExportRevision(
            outcome_set_id=str(outcome_set_id),
            processing_result_id=str(fixture.result_id),
            revision_token=f"{outcome_set_id}:hash:fixture-v1",
            status="available",
            category_states={
                "summary": "available",
                "key_points": "available",
                "decisions": "not_found",
                "action_items": "available",
                "followups": "not_found",
                "risks": "not_found",
                "questions": "not_found",
                "evidence": "available",
            },
            source_kind="extractive_generator",
            generator_kind="deterministic_extractive",
            generator_version="fixture-v1",
            content_hash="fixture-content-hash",
            items=(
                SummaryExportItem(
                    category="summary",
                    sequence=0,
                    state="available",
                    text="Сохранённое саммари.",
                    owner=None,
                    due_date=None,
                    truth_label="supported",
                    source_references=({"transcript_segment_id": raw[0].segment_id},),
                    evidence_turn_ids=(turns[0].turn_id,),
                    unresolved_references=(),
                ),
                SummaryExportItem(
                    category="action_items",
                    sequence=0,
                    state="available",
                    text="@подготовить отчёт",
                    owner="=Иван",
                    due_date="2026-07-30",
                    truth_label="supported",
                    source_references=(),
                    evidence_turn_ids=(),
                    unresolved_references=({"sequence": 999},),
                ),
            ),
        )
    return ExportSnapshot(
        selection=ExportSelection(
            content_scope=scope,  # type: ignore[arg-type]
            format=format,  # type: ignore[arg-type]
            processing_result_id=fixture.result_id,
            outcome_set_id=outcome_set_id,
            include_speaker_labels=include_speaker_labels,
            include_timestamps=include_timestamps,
        ),
        meeting_id="12000000-0000-0000-0000-000000000120",
        meeting_title="Синтетическая встреча <без HTML>",
        language="ru",
        duration_seconds=3702,
        processing_result_id=str(fixture.result_id),
        processing_result_version=3,
        media_revision_id=str(uuid4()),
        raw_segments=raw,
        canonical_turns=turns,
        summary=summary,
    )


def test_canonical_turns_preserve_boundaries_unknown_and_long_gaps(export_fixture) -> None:
    snapshot = _snapshot(export_fixture, format="json")

    assert len(snapshot.raw_segments) == 7
    assert len(snapshot.canonical_turns) == 7
    assert snapshot.raw_segments[3].attribution_state == "uncertain"
    assert snapshot.canonical_turns[3].speaker_label == "Спикер не определён"
    assert snapshot.canonical_turns[0].speaker_label == "Анна"
    assert snapshot.canonical_turns[2].speaker_label == "Анна"
    assert snapshot.canonical_turns[0].turn_id != snapshot.canonical_turns[2].turn_id
    assert all("pause" not in turn.text.lower() for turn in snapshot.canonical_turns)


def test_txt_and_markdown_keep_russian_timestamps_and_escape_markup(export_fixture) -> None:
    txt = render_content_export(_snapshot(export_fixture, format="txt")).body.decode()
    markdown = render_content_export(_snapshot(export_fixture, format="md")).body.decode()

    assert "Привет, GRAF." in txt
    assert "[01:01:40.100]" in txt
    assert "Пауза" not in txt
    assert "&lt;без HTML&gt;" in markdown
    assert "Синтетическая встреча <без HTML>" not in markdown


def test_degraded_state_is_visible_in_human_and_structured_exports(export_fixture) -> None:
    base = _snapshot(export_fixture, format="txt")
    degraded = replace(
        base,
        attribution_result_state="degraded_provider_result",
        attribution_reason_codes=("unknown_tiny_identity",),
        raw_segments=tuple(
            replace(row, result_state="degraded_provider_result") for row in base.raw_segments
        ),
        canonical_turns=tuple(
            replace(turn, result_state="degraded_provider_result")
            for turn in base.canonical_turns
        ),
    )

    text = render_content_export(degraded).body.decode()
    workbook = load_workbook(
        io.BytesIO(render_content_export(replace(degraded, selection=replace(degraded.selection, format="xlsx"))).body),
        read_only=False,
        data_only=False,
    )
    metadata = {row[0].value: row[1].value for row in workbook["Metadata"].iter_rows(min_row=2)}

    assert (
        "Разделение по спикерам: частично готово; фрагменты без имени отмечены как "
        "«Спикер не определён»"
    ) in text
    assert all(row.result_state == "degraded_provider_result" for row in degraded.raw_segments)
    assert metadata["attribution_result_state"] == "degraded_provider_result"
    assert json.loads(metadata["attribution_reason_codes"]) == ["unknown_tiny_identity"]


def test_markdown_neutralizes_line_markers_links_and_raw_html(export_fixture) -> None:
    snapshot = _snapshot(export_fixture, format="md")
    malicious = replace(
        snapshot.canonical_turns[0],
        text="# Заголовок\n- список\n1. пункт\n[ссылка](https://example.test)\n<script>x</script>",
    )
    body = render_content_export(
        replace(snapshot, canonical_turns=(malicious, *snapshot.canonical_turns[1:]))
    ).body.decode()

    assert "\\# Заголовок" in body
    assert "\\- список" in body
    assert "1\\. пункт" in body
    assert "\\[ссылка\\]\\(https\\:\\/\\/example\\.test\\)" in body
    assert "&lt;script&gt;x&lt;\\/script&gt;" in body
    assert "\n# Заголовок" not in body
    assert "\n- список" not in body


def test_csv_has_stable_columns_crlf_bom_and_inert_formula_cells(export_fixture) -> None:
    body = render_content_export(_snapshot(export_fixture, format="csv")).body
    assert body.startswith(b"\xef\xbb\xbf")
    assert b"\r\n" in body
    rows = list(csv.DictReader(io.StringIO(body.decode("utf-8-sig"), newline="")))

    assert tuple(rows[0]) == CSV_COLUMNS
    assert len(rows) == 7
    assert rows[1]["text"].startswith("'=")
    assert json.loads(rows[0]["source_segment_ids"])
    assert rows[3]["attribution_state"] == "unknown"
    assert rows[0]["provider_speaker_key"] == "speaker-a"
    assert rows[0]["result_state"] == "accepted"


def test_json_is_versioned_deterministic_raw_fidelity_without_provider_secrets(
    export_fixture,
) -> None:
    snapshot = _snapshot(export_fixture, format="json")
    first = render_content_export(snapshot).body
    second = render_content_export(snapshot).body
    payload = json.loads(first)

    assert first == second
    assert payload["schema_version"] == SCHEMA_VERSION
    assert payload["provenance"]["turn_policy_version"] == TURN_POLICY_VERSION
    assert len(payload["transcript"]["raw_segments"]) == 7
    assert payload["transcript"]["raw_segments"][3]["text"] == "Неизвестная атрибуция."
    assert "mediascribe_job_id" not in first.decode()
    assert "storage_object_key" not in first.decode()
    assert "api_key" not in first.decode()


def test_provider_speaker_labels_do_not_change_graf_canonical_semantics(export_fixture) -> None:
    adapter_rows = [
        DiarizationSegment(
            id=uuid4(),
            processing_result_id=row.processing_result_id,
            meeting_id=row.meeting_id,
            workspace_id=row.workspace_id,
            sequence=row.sequence,
            start_seconds=row.start_seconds,
            end_seconds=row.end_seconds,
            text=row.text,
            speaker_label={"speaker-a": "remote-7", "speaker-b": "remote-2", "UNKNOWN": "UNKNOWN"}[
                row.speaker_label
            ],
            source_role=row.source_role,
        )
        for row in export_fixture.diarization_rows
    ]
    original_turns = canonical_export_turns(
        list(export_fixture.transcript_rows),
        diarization_rows=list(export_fixture.diarization_rows),
        processing_result_id=export_fixture.result_id,
    )
    adapter_turns = canonical_export_turns(
        list(export_fixture.transcript_rows),
        diarization_rows=adapter_rows,
        processing_result_id=export_fixture.result_id,
    )

    assert [
        turn.canonical_label if hasattr(turn, "canonical_label") else turn.speaker_label
        for turn in original_turns
    ] == [turn.speaker_label for turn in adapter_turns]
    assert [turn.provider_speaker_key for turn in original_turns] != [
        turn.provider_speaker_key for turn in adapter_turns
    ]
    assert [turn.text for turn in original_turns] == [turn.text for turn in adapter_turns]


def test_srt_uses_one_turn_per_cue_preserves_hour_and_has_no_pause(export_fixture) -> None:
    body = render_content_export(_snapshot(export_fixture, format="srt")).body.decode()

    assert body.count(" --> ") == 7
    assert "01:01:40,100 --> 01:01:41,100" in body
    assert "Спикер не определён: Неизвестная атрибуция." in body
    assert "Пауза" not in body


def test_xlsx_has_fixed_safe_sheets_columns_and_literal_cells(export_fixture) -> None:
    body = render_content_export(_snapshot(export_fixture, format="xlsx", scope="combined")).body
    workbook = load_workbook(io.BytesIO(body), read_only=False, data_only=False)

    assert workbook.sheetnames == ["Transcript", "Summary", "Action Items", "Metadata"]
    transcript = workbook["Transcript"]
    assert tuple(cell.value for cell in transcript[1]) == CSV_COLUMNS
    assert transcript.freeze_panes == "A2"
    assert transcript.column_dimensions["M"].width == 72
    assert transcript["M2"].alignment.wrap_text is True
    assert transcript["M3"].value.startswith("'=")
    summary = workbook["Summary"]
    assert [cell.value for cell in summary["A"][1:]] == [
        "summary",
        "key_points",
        "decisions",
        "followups",
        "risks",
        "questions",
        "evidence",
    ]
    assert summary["C3"].value == "available"
    assert summary["C4"].value == "not_found"
    action = workbook["Action Items"]
    assert action["C2"].value.startswith("'@")
    assert action["D2"].value.startswith("'=")
    metadata = {row[0].value: row[1].value for row in workbook["Metadata"].iter_rows(min_row=2)}
    assert metadata["summary_revision_token"]
    assert metadata["summary_source_kind"] == "extractive_generator"
    assert metadata["summary_generator_version"] == "fixture-v1"
    assert json.loads(metadata["summary_category_states"])["summary"] == "available"
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def test_xlsx_marks_unselected_sheets_and_honors_evidence_option(export_fixture) -> None:
    snapshot = _snapshot(export_fixture, format="xlsx", scope="summary")
    snapshot = replace(
        snapshot,
        selection=replace(snapshot.selection, include_evidence=False),
    )
    workbook = load_workbook(
        io.BytesIO(render_content_export(snapshot).body),
        read_only=False,
        data_only=False,
    )

    transcript = workbook["Transcript"]
    assert transcript["B2"].value == "status:not_selected"
    summary = workbook["Summary"]
    assert summary["F2"].value == "[]"
    assert summary["G2"].value == "[]"
    assert summary["H2"].value == "[]"


def test_xlsx_neutralizes_formula_prefixes_in_summary_metadata(export_fixture) -> None:
    snapshot = _snapshot(export_fixture, format="xlsx", scope="summary")
    assert snapshot.summary is not None
    snapshot = replace(
        snapshot,
        summary=replace(snapshot.summary, generator_kind="=UNTRUSTED_METADATA"),
    )
    workbook = load_workbook(
        io.BytesIO(render_content_export(snapshot).body),
        read_only=False,
        data_only=False,
    )
    metadata = {row[0].value: row[1].value for row in workbook["Metadata"].iter_rows(min_row=2)}

    assert metadata["summary_generator_kind"] == "'=UNTRUSTED_METADATA"
    assert not any(cell.data_type == "f" for row in workbook["Metadata"] for cell in row)


def test_summary_only_and_combined_do_not_regenerate_or_invent_fields(export_fixture) -> None:
    summary_json = json.loads(
        render_content_export(_snapshot(export_fixture, format="json", scope="summary")).body
    )
    combined = render_content_export(
        _snapshot(export_fixture, format="txt", scope="combined")
    ).body.decode()

    assert summary_json["transcript"] is None
    assert summary_json["summary"]["generator_version"] == "fixture-v1"
    assert summary_json["summary"]["items"][1]["owner"] == "=Иван"
    assert "Транскрипт" in combined
    assert "Саммари" in combined
    assert "Сохранённое саммари." in combined


def test_legacy_asr_reference_resolves_to_every_overlapping_canonical_turn() -> None:
    first = CanonicalExportTurn(
        turn_id="turn-a",
        sequence=0,
        start_ms=0,
        end_ms=1000,
        text="one",
        speaker_key="speaker-a",
        speaker_label="A",
        attribution_state="confirmed",
        source_role="canonical_mixed",
        source_segment_ids=("provider-a",),
        overlap=False,
    )
    second = replace(
        first,
        turn_id="turn-b",
        sequence=1,
        start_ms=1000,
        end_ms=2000,
        text="two",
        speaker_key="speaker-b",
        speaker_label="B",
        source_segment_ids=("provider-b",),
    )

    resolved = _reference_turn_ids(
        {
            "transcript_segment_id": "legacy-asr",
            "start_seconds": 0,
            "end_seconds": 2,
            "source_role": "mixed",
        },
        (first, second),
        {"provider-a": "turn-a", "provider-b": "turn-b"},
    )

    assert resolved == ("turn-a", "turn-b")


def test_presentation_options_do_not_change_machine_formats(export_fixture) -> None:
    base = _snapshot(export_fixture, format="json")
    toggled = replace(
        base,
        selection=replace(
            base.selection,
            include_speaker_labels=False,
            include_timestamps=False,
        ),
    )

    base_payload = json.loads(render_content_export(base).body)
    toggled_payload = json.loads(render_content_export(toggled).body)
    assert (
        base_payload["transcript"]["canonical_turns"]
        == toggled_payload["transcript"]["canonical_turns"]
    )
    assert toggled_payload["selection"] == {
        "content_scope": "transcript",
        "format": "json",
        "include_evidence": False,
        "include_speaker_labels": True,
        "include_timestamps": True,
    }


def test_human_groups_may_join_short_fragments_but_keep_canonical_children() -> None:
    first = CanonicalExportTurn(
        turn_id="turn-a",
        sequence=0,
        start_ms=0,
        end_ms=1000,
        text="Короткий фрагмент.",
        speaker_key="speaker-a",
        speaker_label="Анна",
        attribution_state="confirmed",
        source_role="local_microphone",
        source_segment_ids=("raw-a",),
        overlap=False,
    )
    second = replace(
        first,
        turn_id="turn-b",
        sequence=1,
        start_ms=4000,
        end_ms=5000,
        text="Второй фрагмент.",
        source_segment_ids=("raw-b",),
    )
    unknown = replace(
        second,
        turn_id="turn-unknown",
        sequence=2,
        start_ms=5000,
        end_ms=5500,
        speaker_key="unknown:raw-c",
        speaker_label="UNKNOWN",
        attribution_state="unknown",
        source_segment_ids=("raw-c",),
    )

    groups = human_display_groups((first, second, unknown))

    assert [[turn.turn_id for turn in group] for group in groups] == [
        ["turn-a", "turn-b"],
        ["turn-unknown"],
    ]


def test_vtt_uses_canonical_turn_boundaries(export_fixture) -> None:
    body = render_content_export(_snapshot(export_fixture, format="vtt")).body.decode()

    assert body.startswith("WEBVTT\n\n")
    assert body.count(" --> ") == 7
    assert "01:01:40.100 --> 01:01:41.100" in body
    assert "Спикер не определён: Неизвестная атрибуция." in body


@pytest.mark.parametrize("format", ["srt", "vtt"])
def test_subtitle_export_fails_closed_for_non_positive_rounded_timing(
    export_fixture,
    format: str,
) -> None:
    snapshot = _snapshot(export_fixture, format=format)
    first, *remaining = snapshot.canonical_turns
    invalid = replace(
        snapshot,
        canonical_turns=(replace(first, end_ms=first.start_ms), *remaining),
    )

    with pytest.raises(ProblemDetail) as exc_info:
        render_content_export(invalid)

    assert exc_info.value.status == 409
    assert exc_info.value.code == "subtitle_timing_unavailable"


def test_zero_duration_raw_evidence_is_marked_invalid(export_fixture) -> None:
    source = export_fixture.transcript_rows[0]
    row = TranscriptSegment(
        id=source.id,
        processing_result_id=source.processing_result_id,
        workspace_id=source.workspace_id,
        meeting_id=source.meeting_id,
        sequence=source.sequence,
        start_seconds=source.start_seconds,
        end_seconds=Decimal(export_fixture.transcript_rows[0].start_seconds),
        text=source.text,
        source_role=source.source_role,
        source_role_original=source.source_role_original,
    )

    segment = canonical_raw_segments([row])[0]

    assert segment.timing_state == "invalid"
    assert segment.omission_reason == "invalid_timing"
