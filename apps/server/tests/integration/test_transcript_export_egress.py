from __future__ import annotations

import asyncio
import io
import json
from datetime import UTC, datetime
from unittest.mock import patch
from uuid import UUID, uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import delete, select

from tests.contract.test_ingest_openapi_contract import auth_headers
from tests.fakes.fake_temporal import FakeTemporalClient
from tests.fixtures.cabinet import SAFE_TRANSCRIPT_TEXT, seed_cabinet_meetings
from tests.fixtures.cabinet_access import (
    add_workspace_user,
    audit_events,
    auth_headers_for,
    grant_meeting_to_user,
    set_artifact_policy,
    set_meeting_deletion_state,
    set_meeting_visibility,
)
from twobrain_rec_server.api.problems import ProblemDetail
from twobrain_rec_server.cabinet import egress as egress_module
from twobrain_rec_server.db.models import (
    DiarizationSegment,
    ExportPackage,
    MediaRevision,
    Meeting,
    MeetingArtifactPolicy,
    MeetingOutcomeItem,
    MeetingOutcomeSet,
    MeetingSummarySlot,
    ProcessingResult,
    TranscriptSegment,
)


def test_implicit_content_policy_is_owner_only_and_explicit_deny_stays_disabled(client) -> None:
    seeds = seed_cabinet_meetings(client)

    owner = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/downloads/transcript",
        headers=auth_headers(),
    )
    assert owner.status_code == 200
    assert SAFE_TRANSCRIPT_TEXT in owner.text

    add_workspace_user(client)
    grant_meeting_to_user(client, seeds.ready_id)
    shared = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/downloads/transcript",
        headers=auth_headers_for(),
    )
    assert shared.status_code == 409

    set_artifact_policy(
        client,
        seeds.ready_id,
        transcript_download="disabled",
        policy_source="meeting_override",
    )
    denied = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/downloads/transcript",
        headers=auth_headers(),
    )
    assert denied.status_code == 409


def test_capability_is_metadata_only_and_separates_transcript_summary_combined(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(
        client,
        seeds.ready_id,
        transcript_download="allowed",
        summary_download="allowed",
    )

    response = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["transcript"] == {"state": "available", "reason": None}
    assert payload["summary"]["state"] == "missing"
    assert payload["combined"]["state"] == "missing"
    assert payload["formats"]["transcript"] == [
        "txt",
        "md",
        "csv",
        "xlsx",
        "json",
        "srt",
        "vtt",
    ]
    assert SAFE_TRANSCRIPT_TEXT not in response.text
    assert "storage_object_key" not in response.text
    assert "mediascribe_job_id" not in response.text


def test_implicit_summary_policy_allows_owner_server_mediated_export(client) -> None:
    seeds = seed_cabinet_meetings(client)
    outcome_set_id = asyncio.run(_seed_stored_summary(client, seeds.ready_id))

    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    )
    assert capability.status_code == 200
    payload = capability.json()
    assert payload["summary"]["state"] == "available"
    assert payload["outcome_set_id"] == str(outcome_set_id)

    exported = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
        json={
            "content_scope": "summary",
            "format": "json",
            "processing_result_id": payload["processing_result_id"],
            "outcome_set_id": payload["outcome_set_id"],
        },
    )
    assert exported.status_code == 200


def test_unaccepted_candidate_never_replaces_the_exported_summary(client) -> None:
    seeds = seed_cabinet_meetings(client)
    accepted_id = asyncio.run(_seed_stored_summary(client, seeds.ready_id))

    async def seed_candidate() -> None:
        async with client.app_state["sessionmaker"]() as db:
            meeting = await db.get(Meeting, seeds.ready_id)
            result = await db.scalar(
                select(ProcessingResult).where(ProcessingResult.meeting_id == seeds.ready_id)
            )
            assert meeting is not None and result is not None
            candidate = MeetingOutcomeSet(
                workspace_id=meeting.workspace_id,
                meeting_id=meeting.id,
                media_revision_id=result.media_revision_id,
                processing_result_id=result.id,
                candidate_id=uuid4(),
                status="available",
                summary_state="available",
                key_points_state="not_found",
                decisions_state="not_found",
                action_items_state="not_found",
                followups_state="not_found",
                risks_state="not_found",
                questions_state="not_found",
                evidence_state="not_found",
                source_kind="litellm",
                generator_kind="litellm",
                generator_version="fixture-private-candidate-v1",
                lifecycle_state="active",
                revision_state="candidate",
                generated_at=datetime.now(UTC),
            )
            db.add(candidate)
            await db.flush()
            db.add(
                MeetingOutcomeItem(
                    workspace_id=meeting.workspace_id,
                    meeting_id=meeting.id,
                    outcome_set_id=candidate.id,
                    category="summary",
                    sequence=0,
                    state="available",
                    text="Непринятый приватный вариант.",
                    truth_label="supported",
                    source_refs_json=[],
                )
            )
            await db.commit()

    asyncio.run(seed_candidate())
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()
    exported = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
        json={
            "content_scope": "summary",
            "format": "txt",
            "processing_result_id": capability["processing_result_id"],
            "outcome_set_id": capability["outcome_set_id"],
        },
    )

    assert capability["outcome_set_id"] == str(accepted_id)
    assert exported.status_code == 200
    assert "Сохранённый итог." in exported.text
    assert "Непринятый приватный вариант." not in exported.text


def test_authorized_transcript_formats_share_one_revision_and_safe_headers(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(client, seeds.ready_id, transcript_download="allowed")
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()
    result_id = capability["processing_result_id"]

    responses = {}
    for format_name in ("txt", "md", "csv", "xlsx", "json", "srt", "vtt"):
        response = client.post(
            f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
            headers=auth_headers(),
            json={
                "content_scope": "transcript",
                "format": format_name,
                "processing_result_id": result_id,
                "outcome_set_id": None,
                "include_speaker_labels": True,
                "include_timestamps": True,
                "include_evidence": True,
            },
        )
        assert response.status_code == 200, (format_name, response.text)
        assert response.headers["x-content-type-options"] == "nosniff"
        assert response.headers["cache-control"] == "private, no-store"
        assert int(response.headers["content-length"]) == len(response.content)
        assert f"-transcript-r1.{format_name}" in response.headers["content-disposition"]
        responses[format_name] = response

    assert SAFE_TRANSCRIPT_TEXT in responses["txt"].text
    assert SAFE_TRANSCRIPT_TEXT.replace(".", "\\.") in responses["md"].text
    assert SAFE_TRANSCRIPT_TEXT in responses["csv"].content.decode("utf-8-sig")
    assert SAFE_TRANSCRIPT_TEXT in responses["srt"].text
    assert SAFE_TRANSCRIPT_TEXT in responses["vtt"].text
    payload = responses["json"].json()
    assert payload["revisions"]["processing_result_id"] == result_id
    assert payload["provenance"]["provider_neutral"] is True
    workbook = load_workbook(io.BytesIO(responses["xlsx"].content), read_only=True)
    assert workbook.sheetnames == ["Transcript", "Summary", "Action Items", "Metadata"]
    events = audit_events(client, seeds.ready_id)
    assert [event.event_type for event in events] == [
        event
        for _ in range(7)
        for event in ("content_export_requested", "content_export_completed")
    ]
    assert all(SAFE_TRANSCRIPT_TEXT not in json.dumps(event.metadata_json) for event in events)


def test_provider_only_legacy_result_stays_hidden_until_transcript_and_diarization_match(client) -> None:
    seeds = seed_cabinet_meetings(client)
    result_id = asyncio.run(_keep_only_provider_turns(client, seeds.ready_id))
    set_artifact_policy(client, seeds.ready_id, transcript_download="allowed")

    page = client.get(f"/meetings/{seeds.ready_id}", headers=auth_headers())
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    )

    assert page.status_code == 200
    assert SAFE_TRANSCRIPT_TEXT not in page.text
    assert capability.status_code == 200
    assert capability.json()["transcript"]["state"] == "missing"

    for format_name in ("txt", "md", "csv", "xlsx", "json", "srt", "vtt"):
        response = client.post(
            f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
            headers=auth_headers(),
            json={
                "content_scope": "transcript",
                "format": format_name,
                "processing_result_id": str(result_id),
            },
        )
        assert response.status_code == 409, (format_name, response.text)


def test_summary_and_combined_use_current_stored_outcome_without_regeneration(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(
        client,
        seeds.ready_id,
        transcript_download="allowed",
        summary_download="allowed",
    )
    outcome_set_id = asyncio.run(_seed_stored_summary(client, seeds.ready_id))
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()

    assert capability["outcome_set_id"] == str(outcome_set_id)
    assert capability["summary"]["state"] == "available"
    assert capability["combined"]["state"] == "available"
    for scope in ("summary", "combined"):
        response = client.post(
            f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
            headers=auth_headers(),
            json={
                "content_scope": scope,
                "format": "json",
                "processing_result_id": capability["processing_result_id"],
                "outcome_set_id": capability["outcome_set_id"],
            },
        )
        assert response.status_code == 200, response.text
        payload = response.json()
        assert payload["summary"]["outcome_set_id"] == str(outcome_set_id)
        assert payload["summary"]["items"][0]["text"] == "Сохранённый итог."
        assert [item["category"] for item in payload["summary"]["items"]] == [
            "summary",
            "decisions",
            "action_items",
        ]
        assert (payload["transcript"] is None) == (scope == "summary")


@pytest.mark.parametrize(
    ("transcript_policy", "summary_policy", "transcript_state", "summary_state", "combined_state"),
    [
        ("allowed", "allowed", "available", "available", "available"),
        ("allowed", "disabled", "available", "denied", "missing"),
        ("disabled", "allowed", "denied", "available", "missing"),
        ("disabled", "disabled", "denied", "denied", "missing"),
    ],
)
def test_combined_policy_is_composed_fail_closed_from_component_policies(
    client,
    transcript_policy: str,
    summary_policy: str,
    transcript_state: str,
    summary_state: str,
    combined_state: str,
) -> None:
    seeds = seed_cabinet_meetings(client)
    _ = asyncio.run(_seed_stored_summary(client, seeds.ready_id))
    set_artifact_policy(
        client,
        seeds.ready_id,
        transcript_download=transcript_policy,
        summary_download=summary_policy,
        package_export="allowed",
    )

    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()

    assert capability["transcript"]["state"] == transcript_state
    assert capability["summary"]["state"] == summary_state
    assert capability["combined"]["state"] == combined_state


@pytest.mark.parametrize(
    ("outcome_status", "summary_state", "combined_state"),
    [
        ("partial", "partial", "available"),
        ("generating", "missing", "missing"),
        ("failed", "missing", "missing"),
    ],
)
def test_summary_capability_preserves_stored_partial_processing_and_failed_truth(
    client,
    outcome_status: str,
    summary_state: str,
    combined_state: str,
) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(
        client,
        seeds.ready_id,
        transcript_download="allowed",
        summary_download="allowed",
    )
    outcome_set_id = asyncio.run(
        _seed_stored_summary(client, seeds.ready_id, status=outcome_status)
    )

    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    )

    assert capability.status_code == 200
    payload = capability.json()
    if outcome_status == "partial":
        assert payload["outcome_set_id"] == str(outcome_set_id)
    else:
        assert payload["outcome_set_id"] is None
    assert payload["summary"]["state"] == summary_state
    assert payload["combined"]["state"] == combined_state
    if outcome_status == "partial":
        exported = client.post(
            f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
            headers=auth_headers(),
            json={
                "content_scope": "summary",
                "format": "json",
                "processing_result_id": payload["processing_result_id"],
                "outcome_set_id": payload["outcome_set_id"],
            },
        )
        assert exported.status_code == 200
        assert exported.json()["summary"]["status"] == "partial"


def test_unsupported_scope_format_stale_revision_and_policy_fail_closed(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(client, seeds.ready_id, transcript_download="allowed")
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()
    path = f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports"

    incompatible = client.post(
        path,
        headers=auth_headers(),
        json={
            "content_scope": "transcript",
            "format": "pdf",
            "processing_result_id": capability["processing_result_id"],
        },
    )
    stale = client.post(
        path,
        headers=auth_headers(),
        json={
            "content_scope": "transcript",
            "format": "txt",
            "processing_result_id": str(uuid4()),
        },
    )
    asyncio.run(_update_transcript_policy(client, seeds.ready_id, "disabled"))
    denied = client.post(
        path,
        headers=auth_headers(),
        json={
            "content_scope": "transcript",
            "format": "txt",
            "processing_result_id": capability["processing_result_id"],
        },
    )

    assert incompatible.status_code == 422
    assert stale.status_code == 409
    assert stale.json()["code"] == "export_revision_stale"
    assert denied.status_code == 409
    assert denied.json()["code"] == "export_unavailable"


def test_deletion_in_progress_blocks_capability_and_file_without_serialized_bytes(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(client, seeds.ready_id, transcript_download="allowed")
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()
    set_meeting_deletion_state(client, seeds.ready_id, "requested")

    blocked_capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    )
    blocked_file = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
        json={
            "content_scope": "transcript",
            "format": "json",
            "processing_result_id": capability["processing_result_id"],
        },
    )

    assert blocked_capability.status_code == 200
    assert blocked_capability.json()["transcript"]["state"] == "deletion_in_progress"
    assert blocked_file.status_code == 409
    assert blocked_file.json()["code"] == "meeting_deletion_active"
    assert SAFE_TRANSCRIPT_TEXT not in blocked_file.text
    assert [
        (event.event_type, event.outcome, event.policy_reason)
        for event in audit_events(client, seeds.ready_id)
    ] == [("content_export_denied", "denied", "meeting_deletion_active")]


def test_deletion_started_during_render_is_rechecked_before_bytes_escape(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(client, seeds.ready_id, transcript_download="allowed")
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()
    original_build = egress_module.build_export_snapshot

    async def build_then_start_deletion(*args, **kwargs):
        snapshot = await original_build(*args, **kwargs)
        async with client.app_state["sessionmaker"]() as db:
            meeting = await db.get(Meeting, seeds.ready_id)
            assert meeting is not None
            meeting.deletion_state = "requested"
            await db.commit()
        return snapshot

    with patch.object(egress_module, "build_export_snapshot", build_then_start_deletion):
        response = client.post(
            f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
            headers=auth_headers(),
            json={
                "content_scope": "transcript",
                "format": "txt",
                "processing_result_id": capability["processing_result_id"],
            },
        )

    assert response.status_code == 409
    assert response.json()["code"] == "meeting_deletion_active"
    assert SAFE_TRANSCRIPT_TEXT not in response.text
    assert "content-disposition" not in response.headers
    assert [event.event_type for event in audit_events(client, seeds.ready_id)] == [
        "content_export_requested",
        "content_export_denied",
    ]


def test_unexpected_snapshot_failure_returns_safe_generation_error_and_audits(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(client, seeds.ready_id, transcript_download="allowed")
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()

    async def fail_snapshot(*args, **kwargs):
        raise RuntimeError("synthetic snapshot failure")

    with patch.object(egress_module, "build_export_snapshot", fail_snapshot):
        response = client.post(
            f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
            headers=auth_headers(),
            json={
                "content_scope": "transcript",
                "format": "txt",
                "processing_result_id": capability["processing_result_id"],
            },
        )

    assert response.status_code == 503
    assert response.json()["code"] == "export_generation_failed"
    assert SAFE_TRANSCRIPT_TEXT not in response.text
    assert "content-disposition" not in response.headers
    assert [
        (event.event_type, event.outcome, event.policy_reason)
        for event in audit_events(client, seeds.ready_id)
    ] == [
        ("content_export_requested", "allowed", "policy_allowed"),
        ("content_export_failed", "failed", "export_generation_failed"),
    ]


def test_access_revoked_after_capability_read_returns_no_attachment(client) -> None:
    seeds = seed_cabinet_meetings(client)
    add_workspace_user(client)
    set_meeting_visibility(client, seeds.ready_id, "team_visible")
    set_artifact_policy(client, seeds.ready_id, transcript_download="allowed")
    headers = auth_headers_for()
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=headers,
    )
    assert capability.status_code == 200
    set_meeting_visibility(client, seeds.ready_id, "owner_only")

    response = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=headers,
        json={
            "content_scope": "transcript",
            "format": "txt",
            "processing_result_id": capability.json()["processing_result_id"],
        },
    )

    assert response.status_code == 404
    assert SAFE_TRANSCRIPT_TEXT not in response.text
    assert "content-disposition" not in response.headers


def test_owner_team_shared_and_denied_access_states_are_server_enforced(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(client, seeds.ready_id, transcript_download="allowed")
    owner = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    )
    assert owner.status_code == 200
    assert owner.json()["transcript"]["state"] == "available"

    add_workspace_user(client)
    set_meeting_visibility(client, seeds.ready_id, "team_visible")
    team = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers_for(),
    )
    assert team.status_code == 200
    assert team.json()["transcript"]["state"] == "available"

    set_meeting_visibility(client, seeds.ready_id, "owner_only")
    grant_meeting_to_user(client, seeds.ready_id)
    shared = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers_for(),
    )
    assert shared.status_code == 200
    assert shared.json()["transcript"]["state"] == "available"

    denied_user_id = UUID("30000000-0000-0000-0000-000000000120")
    denied_device_id = UUID("40000000-0000-0000-0000-000000000120")
    add_workspace_user(
        client,
        user_id=denied_user_id,
        device_id=denied_device_id,
        display_name="Denied User",
    )
    denied = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers_for(user_id=denied_user_id, device_id=denied_device_id),
    )
    assert denied.status_code == 404


def test_access_revoked_during_render_returns_no_attachment(client) -> None:
    seeds = seed_cabinet_meetings(client)
    add_workspace_user(client)
    set_meeting_visibility(client, seeds.ready_id, "team_visible")
    set_artifact_policy(client, seeds.ready_id, transcript_download="allowed")
    headers = auth_headers_for()
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=headers,
    ).json()
    original_build = egress_module.build_export_snapshot

    async def build_then_revoke_access(*args, **kwargs):
        snapshot = await original_build(*args, **kwargs)
        async with client.app_state["sessionmaker"]() as db:
            meeting = await db.get(Meeting, seeds.ready_id)
            assert meeting is not None
            meeting.visibility = "owner_only"
            await db.commit()
        return snapshot

    with patch.object(egress_module, "build_export_snapshot", build_then_revoke_access):
        response = client.post(
            f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
            headers=headers,
            json={
                "content_scope": "transcript",
                "format": "txt",
                "processing_result_id": capability["processing_result_id"],
            },
        )

    assert response.status_code == 404
    assert response.json()["code"] == "meeting_not_found"
    assert SAFE_TRANSCRIPT_TEXT not in response.text
    assert "content-disposition" not in response.headers
    assert [event.event_type for event in audit_events(client, seeds.ready_id)] == [
        "content_export_requested",
        "content_export_denied",
    ]


def test_policy_revoked_during_render_returns_no_attachment(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(client, seeds.ready_id, transcript_download="allowed")
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()
    original_build = egress_module.build_export_snapshot

    async def build_then_revoke_policy(*args, **kwargs):
        snapshot = await original_build(*args, **kwargs)
        await _update_transcript_policy(client, seeds.ready_id, "disabled")
        return snapshot

    with patch.object(egress_module, "build_export_snapshot", build_then_revoke_policy):
        response = client.post(
            f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
            headers=auth_headers(),
            json={
                "content_scope": "transcript",
                "format": "txt",
                "processing_result_id": capability["processing_result_id"],
            },
        )

    assert response.status_code == 403
    assert response.json()["code"] == "export_policy_denied"
    assert SAFE_TRANSCRIPT_TEXT not in response.text
    assert "content-disposition" not in response.headers


def test_new_generation_does_not_hide_current_saved_summary(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(
        client,
        seeds.ready_id,
        transcript_download="allowed",
        summary_download="allowed",
    )
    saved_id = asyncio.run(
        _seed_stored_summary(client, seeds.ready_id, generator_version="fixture-saved-v1")
    )
    generating_id = asyncio.run(
        _seed_stored_summary(
            client,
            seeds.ready_id,
            generator_version="fixture-generating-v2",
            status="generating",
        )
    )

    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()

    assert generating_id != saved_id
    assert capability["outcome_set_id"] == str(saved_id)
    assert capability["summary"]["state"] == "available"
    assert capability["combined"]["state"] == "available"


def test_export_capability_never_pairs_an_accepted_summary_with_a_newer_result(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(
        client,
        seeds.ready_id,
        transcript_download="allowed",
        summary_download="allowed",
    )
    accepted_id = asyncio.run(_seed_stored_summary(client, seeds.ready_id))

    async def seed_newer_result():
        async with client.app_state["sessionmaker"]() as db:
            current = await db.scalar(
                select(ProcessingResult).where(ProcessingResult.meeting_id == seeds.ready_id)
            )
            assert current is not None
            current_segments = (
                await db.scalars(
                    select(TranscriptSegment)
                    .where(TranscriptSegment.processing_result_id == current.id)
                    .order_by(TranscriptSegment.sequence.asc())
                )
            ).all()
            current_diarization = (
                await db.scalars(
                    select(DiarizationSegment)
                    .where(DiarizationSegment.processing_result_id == current.id)
                    .order_by(DiarizationSegment.sequence.asc())
                )
            ).all()
            newer = ProcessingResult(
                meeting_id=current.meeting_id,
                media_revision_id=current.media_revision_id,
                workspace_id=current.workspace_id,
                mediascribe_job_id=current.mediascribe_job_id,
                processing_workflow_id=current.processing_workflow_id,
                result_version=current.result_version + 1,
                status="imported",
                transcript_status="available",
                diarization_status=current.diarization_status,
                summary_status=current.summary_status,
                language=current.language,
                segment_count=current.segment_count,
                diarization_segment_count=current.diarization_segment_count,
                source_result_hash="newer-result-hash",
                imported_at=datetime.now(UTC),
            )
            db.add(newer)
            await db.flush()
            db.add_all(
                [
                    TranscriptSegment(
                        processing_result_id=newer.id,
                        meeting_id=segment.meeting_id,
                        workspace_id=segment.workspace_id,
                        sequence=segment.sequence,
                        start_seconds=segment.start_seconds,
                        end_seconds=segment.end_seconds,
                        text=segment.text,
                        source_role=segment.source_role,
                        source_role_original=segment.source_role_original,
                    )
                    for segment in current_segments
                ]
            )
            db.add_all(
                [
                    DiarizationSegment(
                        processing_result_id=newer.id,
                        meeting_id=segment.meeting_id,
                        workspace_id=segment.workspace_id,
                        sequence=segment.sequence,
                        start_seconds=segment.start_seconds,
                        end_seconds=segment.end_seconds,
                        text=segment.text,
                        source_role=segment.source_role,
                        speaker_label=segment.speaker_label,
                    )
                    for segment in current_diarization
                ]
            )
            await db.commit()
            return newer.id

    newer_result_id = asyncio.run(seed_newer_result())
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()

    assert capability["processing_result_id"] == str(newer_result_id)
    assert capability["transcript"]["state"] == "available"
    assert capability["summary"] == {
        "state": "available",
        "reason": "stored_summary_revision_stale",
    }
    assert capability["combined"]["state"] == "missing"
    assert capability["outcome_set_id"] == str(accepted_id)

    page = client.get(f"/meetings/{seeds.ready_id}", headers=auth_headers())
    assert page.status_code == 200
    assert f'data-current-outcome-set-id="{accepted_id}"' in page.text

    client.app.state.settings.outcome_generation_enabled = True
    temporal = FakeTemporalClient()
    client.app.state.outcome_temporal_client = temporal
    candidate = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/summary-candidates",
        headers=auth_headers(),
        json={
            "template_key": "graf-auto-v1",
            "template_id": None,
            "template_version": 1,
            "expected_current_outcome_set_id": str(accepted_id),
        },
    )
    assert candidate.status_code == 202, candidate.text
    assert candidate.json()["current_outcome_set_id"] == str(accepted_id)
    assert len(temporal.starts) == 1

    summary_export = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
        json={
            "content_scope": "summary",
            "format": "json",
            "processing_result_id": str(newer_result_id),
            "outcome_set_id": str(accepted_id),
        },
    )
    assert summary_export.status_code == 200
    exported_payload = summary_export.json()
    summary_payload = exported_payload["summary"]
    assert summary_payload["processing_result_id"] != str(newer_result_id)
    assert (
        exported_payload["revisions"]["processing_result_id"]
        == summary_payload["processing_result_id"]
    )
    assert all(not item["evidence_turn_ids"] for item in summary_payload["items"])


def test_export_capability_uses_newest_media_revision_before_summary_acceptance(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(
        client,
        seeds.ready_id,
        transcript_download="allowed",
        summary_download="allowed",
    )
    accepted_id = asyncio.run(_seed_stored_summary(client, seeds.ready_id))

    async def seed_newer_revision_result() -> UUID:
        async with client.app_state["sessionmaker"]() as db:
            current = await db.scalar(
                select(ProcessingResult).where(ProcessingResult.meeting_id == seeds.ready_id)
            )
            meeting = await db.get(Meeting, seeds.ready_id)
            assert current is not None and meeting is not None
            current_segments = (
                await db.scalars(
                    select(TranscriptSegment)
                    .where(TranscriptSegment.processing_result_id == current.id)
                    .order_by(TranscriptSegment.sequence.asc())
                )
            ).all()
            current_diarization = (
                await db.scalars(
                    select(DiarizationSegment)
                    .where(DiarizationSegment.processing_result_id == current.id)
                    .order_by(DiarizationSegment.sequence.asc())
                )
            ).all()
            revision = MediaRevision(
                workspace_id=meeting.workspace_id,
                meeting_id=meeting.id,
                local_media_revision_id="export-newer-media-revision",
                revision_number=2,
                source_kind="reprocess",
                status="accepted",
                manifest_sha256="d" * 64,
                track_sha256_by_role={"media": "e" * 64},
                duration_seconds=meeting.duration_seconds,
                immutable=True,
                accepted_at=datetime.now(UTC),
            )
            db.add(revision)
            await db.flush()
            newer = ProcessingResult(
                meeting_id=current.meeting_id,
                media_revision_id=revision.id,
                workspace_id=current.workspace_id,
                mediascribe_job_id=current.mediascribe_job_id,
                processing_workflow_id=current.processing_workflow_id,
                result_version=current.result_version + 1,
                status="imported",
                transcript_status="available",
                diarization_status=current.diarization_status,
                summary_status=current.summary_status,
                language=current.language,
                segment_count=current.segment_count,
                diarization_segment_count=current.diarization_segment_count,
                source_result_hash="newer-media-revision-result-hash",
                imported_at=datetime.now(UTC),
            )
            db.add(newer)
            await db.flush()
            db.add_all(
                [
                    TranscriptSegment(
                        processing_result_id=newer.id,
                        meeting_id=segment.meeting_id,
                        workspace_id=segment.workspace_id,
                        sequence=segment.sequence,
                        start_seconds=segment.start_seconds,
                        end_seconds=segment.end_seconds,
                        text=segment.text,
                        source_role=segment.source_role,
                        source_role_original=segment.source_role_original,
                    )
                    for segment in current_segments
                ]
            )
            db.add_all(
                [
                    DiarizationSegment(
                        processing_result_id=newer.id,
                        meeting_id=segment.meeting_id,
                        workspace_id=segment.workspace_id,
                        sequence=segment.sequence,
                        start_seconds=segment.start_seconds,
                        end_seconds=segment.end_seconds,
                        text=segment.text,
                        source_role=segment.source_role,
                        speaker_label=segment.speaker_label,
                    )
                    for segment in current_diarization
                ]
            )
            await db.commit()
            return newer.id

    newer_result_id = asyncio.run(seed_newer_revision_result())
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    )

    assert capability.status_code == 200
    payload = capability.json()
    assert payload["processing_result_id"] == str(newer_result_id)
    assert payload["outcome_set_id"] == str(accepted_id)
    assert payload["summary"] == {
        "state": "available",
        "reason": "stored_summary_revision_stale",
    }
    transcript_export = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
        json={
            "content_scope": "transcript",
            "format": "txt",
            "processing_result_id": str(newer_result_id),
            "outcome_set_id": None,
        },
    )
    assert transcript_export.status_code == 200, transcript_export.text
    assert SAFE_TRANSCRIPT_TEXT in transcript_export.text


def test_summary_without_content_hash_is_not_exportable_as_a_pinned_revision(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(
        client,
        seeds.ready_id,
        transcript_download="allowed",
        summary_download="allowed",
    )
    outcome_id = asyncio.run(_seed_stored_summary(client, seeds.ready_id))

    async def clear_content_hash() -> None:
        async with client.app_state["sessionmaker"]() as db:
            outcome = await db.get(MeetingOutcomeSet, outcome_id)
            assert outcome is not None
            outcome.content_hash = None
            await db.commit()

    asyncio.run(clear_content_hash())
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()

    assert capability["summary"] == {
        "state": "failed",
        "reason": "stored_summary_revision_unpinned",
    }
    assert capability["combined"]["state"] == "missing"


def test_requested_and_completion_audit_failures_return_no_attachment(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(client, seeds.ready_id, transcript_download="allowed")
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()
    payload = {
        "content_scope": "transcript",
        "format": "txt",
        "processing_result_id": capability["processing_result_id"],
    }
    audit_problem = ProblemDetail(
        status=503,
        code="audit_unavailable",
        title="Audit unavailable",
    )

    async def fail_requested(*_args, **_kwargs):
        raise audit_problem

    with patch.object(egress_module, "record_egress_audit_event", fail_requested):
        requested = client.post(
            f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
            headers=auth_headers(),
            json=payload,
        )

    original = egress_module.record_egress_audit_event
    calls = 0

    async def fail_completion(*args, **kwargs):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise audit_problem
        return await original(*args, **kwargs)

    with patch.object(egress_module, "record_egress_audit_event", fail_completion):
        completion = client.post(
            f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
            headers=auth_headers(),
            json=payload,
        )

    for response in (requested, completion):
        assert response.status_code == 503
        assert response.json()["code"] == "audit_unavailable"
        assert SAFE_TRANSCRIPT_TEXT not in response.text
        assert "content-disposition" not in response.headers


def test_newer_summary_revision_after_capability_fails_closed_without_mixed_bytes(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(
        client,
        seeds.ready_id,
        transcript_download="allowed",
        summary_download="allowed",
    )
    old_outcome_id = asyncio.run(
        _seed_stored_summary(client, seeds.ready_id, generator_version="fixture-export-v1")
    )
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()
    assert capability["outcome_set_id"] == str(old_outcome_id)
    newer_outcome_id = asyncio.run(
        _seed_stored_summary(client, seeds.ready_id, generator_version="fixture-export-v2")
    )
    assert newer_outcome_id != old_outcome_id

    response = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
        json={
            "content_scope": "summary",
            "format": "json",
            "processing_result_id": capability["processing_result_id"],
            "outcome_set_id": capability["outcome_set_id"],
        },
    )

    assert response.status_code == 409
    assert response.json()["code"] == "export_revision_stale"
    assert "Сохранённый итог." not in response.text
    assert "content-disposition" not in response.headers
    assert [event.event_type for event in audit_events(client, seeds.ready_id)] == [
        "content_export_requested",
        "content_export_denied",
    ]


def test_same_summary_id_with_changed_content_hash_fails_closed(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(
        client,
        seeds.ready_id,
        transcript_download="allowed",
        summary_download="allowed",
    )
    outcome_id = asyncio.run(_seed_stored_summary(client, seeds.ready_id))
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()
    original_build = egress_module.build_export_snapshot

    async def build_then_replace_summary_hash(*args, **kwargs):
        snapshot = await original_build(*args, **kwargs)
        async with client.app_state["sessionmaker"]() as db:
            outcome = await db.get(MeetingOutcomeSet, outcome_id)
            assert outcome is not None
            outcome.content_hash = "synthetic-replaced-content-hash"
            await db.commit()
        return snapshot

    with patch.object(
        egress_module,
        "build_export_snapshot",
        build_then_replace_summary_hash,
    ):
        response = client.post(
            f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
            headers=auth_headers(),
            json={
                "content_scope": "summary",
                "format": "json",
                "processing_result_id": capability["processing_result_id"],
                "outcome_set_id": capability["outcome_set_id"],
            },
        )

    assert response.status_code == 409
    assert response.json()["code"] == "export_revision_stale"
    assert "Сохранённый итог." not in response.text
    assert "content-disposition" not in response.headers


def test_identical_json_retry_returns_identical_revision_pinned_bytes(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(client, seeds.ready_id, transcript_download="allowed")
    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()
    payload = {
        "content_scope": "transcript",
        "format": "json",
        "processing_result_id": capability["processing_result_id"],
    }

    first = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
        json=payload,
    )
    second = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
        json=payload,
    )

    assert first.status_code == second.status_code == 200
    assert first.content == second.content
    assert first.headers["content-disposition"] == second.headers["content-disposition"]


def test_summary_export_pins_persisted_default_and_survives_same_type_refresh(client) -> None:
    seeds = seed_cabinet_meetings(client)
    set_artifact_policy(client, seeds.ready_id, summary_download="allowed")
    first_id = asyncio.run(
        _seed_stored_summary(client, seeds.ready_id, generator_version="fixture-default-v1")
    )

    capability = client.get(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
    ).json()
    first = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
        json={
            "content_scope": "summary",
            "format": "json",
            "processing_result_id": capability["processing_result_id"],
            "outcome_set_id": str(first_id),
        },
    )
    assert first.status_code == 200, first.text

    async def refresh_default() -> UUID:
        second_id = await _seed_stored_summary(
            client, seeds.ready_id, generator_version="fixture-default-v2"
        )
        async with client.app_state["sessionmaker"]() as db:
            slot = await db.scalar(
                select(MeetingSummarySlot).where(
                    MeetingSummarySlot.meeting_id == seeds.ready_id,
                    MeetingSummarySlot.is_meeting_default.is_(True),
                )
            )
            assert slot is not None
            slot.current_outcome_set_id = second_id
            slot.current_binding_class = "verified_complete"
            slot.legacy_migration_proof_hash = None
            await db.commit()
        return second_id

    second_id = asyncio.run(refresh_default())
    assert second_id != first_id

    # The old revision is no longer the default and cannot be exported by UUID.
    stale = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/content-exports",
        headers=auth_headers(),
        json={
            "content_scope": "summary",
            "format": "json",
            "processing_result_id": capability["processing_result_id"],
            "outcome_set_id": str(first_id),
        },
    )
    assert stale.status_code == 409


def test_summary_export_package_manifest_pins_default_revision(client) -> None:
    seeds = seed_cabinet_meetings(client)
    asyncio.run(_seed_stored_summary(client, seeds.ready_id, generator_version="fixture-package-v1"))
    set_artifact_policy(client, seeds.ready_id, package_export="allowed", summary_download="allowed")

    created = client.post(
        f"/api/v1/cabinet/meetings/{seeds.ready_id}/exports",
        headers=auth_headers(),
        json={"artifact_classes": ["summary"]},
    )
    assert created.status_code == 202, created.text

    async def read_manifest() -> dict:
        async with client.app_state["sessionmaker"]() as db:
            package = await db.get(ExportPackage, UUID(created.json()["export_id"]))
            assert package is not None
            return package.manifest_json

    manifest = asyncio.run(read_manifest())
    assert manifest["summary_revision"]["template_key"] == "graf-auto-v1"
    assert manifest["summary_revision"]["outcome_set_id"]


async def _seed_stored_summary(
    client,
    meeting_id: UUID,
    *,
    generator_version: str = "fixture-export-v1",
    status: str = "available",
) -> UUID:
    async with client.app_state["sessionmaker"]() as db:
        result = await db.scalar(
            select(ProcessingResult).where(
                ProcessingResult.meeting_id == meeting_id,
                ProcessingResult.status == "imported",
            )
        )
        meeting = await db.get(Meeting, meeting_id)
        segment = await db.scalar(
            select(TranscriptSegment)
            .where(TranscriptSegment.meeting_id == meeting_id)
            .order_by(TranscriptSegment.sequence.asc())
        )
        assert result is not None and meeting is not None and segment is not None
        outcome_set = MeetingOutcomeSet(
            workspace_id=meeting.workspace_id,
            meeting_id=meeting_id,
            media_revision_id=result.media_revision_id,
            processing_result_id=result.id,
            status=status,
            summary_state=status,
            key_points_state="not_found",
            decisions_state="available",
            action_items_state="available",
            followups_state="not_found",
            risks_state="not_found",
            questions_state="not_found",
            evidence_state="available",
            source_kind="extractive_generator",
            generator_kind="deterministic_extractive",
            generator_version=generator_version,
            template_key="graf-auto-v1",
            template_version=1,
            content_hash=f"fixture-summary-hash-{generator_version}",
            lifecycle_state="active",
            generated_at=datetime.now(UTC) if status in {"available", "partial"} else None,
            revision_state="accepted" if status in {"available", "partial"} else None,
        )
        db.add(outcome_set)
        await db.flush()
        if status in {"available", "partial"}:
            outcome_set.accepted_at = outcome_set.generated_at
            meeting.current_outcome_set_id = outcome_set.id
            slot = await db.scalar(
                select(MeetingSummarySlot).where(
                    MeetingSummarySlot.meeting_id == meeting_id,
                    MeetingSummarySlot.template_key == "graf-auto-v1",
                )
            )
            if slot is None:
                slot = MeetingSummarySlot(
                    workspace_id=meeting.workspace_id,
                    meeting_id=meeting_id,
                    template_key="graf-auto-v1",
                    is_meeting_default=True,
                    default_resolution_source="explicit_meeting",
                    default_resolution_version="slot-fixture-v1",
                    default_resolved_at=datetime.now(UTC),
                )
                db.add(slot)
                await db.flush()
            slot.current_outcome_set_id = outcome_set.id
            slot.current_binding_class = "verified_complete"
        db.add_all(
            [
                MeetingOutcomeItem(
                    workspace_id=meeting.workspace_id,
                    meeting_id=meeting_id,
                    outcome_set_id=outcome_set.id,
                    category="action_items",
                    sequence=0,
                    state="available",
                    text="Сохранённая задача.",
                    truth_label="supported",
                    source_refs_json=[],
                ),
                MeetingOutcomeItem(
                    workspace_id=meeting.workspace_id,
                    meeting_id=meeting_id,
                    outcome_set_id=outcome_set.id,
                    category="decisions",
                    sequence=0,
                    state="available",
                    text="Сохранённое решение.",
                    truth_label="supported",
                    source_refs_json=[],
                ),
                MeetingOutcomeItem(
                    workspace_id=meeting.workspace_id,
                    meeting_id=meeting_id,
                    outcome_set_id=outcome_set.id,
                    category="summary",
                    sequence=0,
                    state="available",
                    text="Сохранённый итог.",
                    truth_label="supported",
                    source_refs_json=[
                        {
                            "transcript_segment_id": str(segment.id),
                            "sequence": segment.sequence,
                            "start_seconds": float(segment.start_seconds),
                            "end_seconds": float(segment.end_seconds),
                            "evidence_kind": "segment",
                        }
                    ],
                ),
            ]
        )
        await db.commit()
        return outcome_set.id


async def _keep_only_provider_turns(client, meeting_id: UUID) -> UUID:
    async with client.app_state["sessionmaker"]() as db:
        result = await db.scalar(
            select(ProcessingResult).where(ProcessingResult.meeting_id == meeting_id)
        )
        assert result is not None
        assert await db.scalar(
            select(DiarizationSegment.id).where(
                DiarizationSegment.processing_result_id == result.id
            )
        )
        await db.execute(
            delete(TranscriptSegment).where(
                TranscriptSegment.processing_result_id == result.id
            )
        )
        result.segment_count = 0
        await db.commit()
        return result.id


async def _update_transcript_policy(client, meeting_id: UUID, value: str) -> None:
    async with client.app_state["sessionmaker"]() as db:
        policy = await db.scalar(
            select(MeetingArtifactPolicy).where(MeetingArtifactPolicy.meeting_id == meeting_id)
        )
        assert policy is not None
        policy.transcript_download = value
        await db.commit()
